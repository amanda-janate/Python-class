'''Automatização de tarefas com excel'''
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = load_workbook(WORKBOOK_PATH)

name = 'Py-Sheet'

worksheet: Worksheet = workbook[name]  # type: ignore

for row in worksheet.iter_rows(min_row=2):
    for col in row:
        print(col.value, end='\t')

        if col.value == 'Maria':
            worksheet.cell(col.row, 2, 23)   # type: ignore
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)